#!/usr/bin/env python3
"""
Update post history Excel file with today's 2 posts (2-post daily model).

Reads linkedin_posts_YYYYMMDD.txt, extracts post metadata, and appends to post_history.xlsx.
Each row: Date, Stream, Post Text (truncated to 500 chars), Source URL, Word Count, Format Used, Notes.
"""

import os
import json
import datetime
import openpyxl
import re

def count_words(text):
    """Count words in a text string."""
    return len(text.split())

def extract_posts_from_file(file_path):
    """
    Parse linkedin_posts_YYYYMMDD.txt and extract 2 posts (News + rotating type).
    Returns list of dicts: [{'stream': '...', 'text': '...', 'source': '...', 'word_count': N, 'format': '...'}, ...]
    """
    if not os.path.exists(file_path):
        print(f"Error: {file_path} not found")
        return []
    
    with open(file_path, 'r', encoding='utf-8') as f:
        content = f.read()
    
    posts = []
    
    # Split by stream markers: STREAM 1 — NEWS, STREAM 2 — [TYPE]
    stream_pattern = r'==+\s*STREAM (\d+) — (.+?)\s*==+'
    stream_blocks = re.split(stream_pattern, content)
    
    # stream_blocks will be: ['prefix', '1', 'NEWS', 'post_text', '2', 'TIPS & TRICKS', 'post_text', ...]
    for i in range(1, len(stream_blocks), 3):
        if i + 2 < len(stream_blocks):
            stream_num = stream_blocks[i].strip()
            stream_name = stream_blocks[i+1].strip()
            post_content = stream_blocks[i+2].strip()
            
            # Extract source URL, word count, format from metadata at end of each stream block
            source_match = re.search(r'Source:\s*(.+?)(?:\n|$)', post_content)
            source_url = source_match.group(1).strip() if source_match else ""
            
            word_count_match = re.search(r'Word count:\s*(\d+)', post_content)
            word_count = int(word_count_match.group(1)) if word_count_match else count_words(post_content)
            
            format_match = re.search(r'Format:\s*(.+?)(?:\n|$)', post_content)
            format_used = format_match.group(1).strip() if format_match else ""
            
            # Extract main post text (remove metadata lines at end)
            post_text = re.sub(r'\n(Source:|Word count:|Format:).*', '', post_content).strip()
            post_text_truncated = post_text[:500] if len(post_text) > 500 else post_text
            
            posts.append({
                'stream_num': int(stream_num),
                'stream_name': stream_name,
                'post_text': post_text_truncated,
                'full_text': post_text,
                'source_url': source_url,
                'word_count': word_count,
                'format_used': format_used
            })
    
    return sorted(posts, key=lambda x: x['stream_num'])

def update_excel_history(posts, output_file='post_history.xlsx'):
    """
    Append posts to post_history.xlsx (2 posts per day in 2-post model).
    Columns: Date | Stream | Post Text (500 char preview) | Full Text | Source URL | Word Count | Format | Notes
    """
    today = datetime.date.today().isoformat()
    
    # Create or load workbook
    if os.path.exists(output_file):
        wb = openpyxl.load_workbook(output_file)
        ws = wb.active
        next_row = ws.max_row + 1
    else:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Post History"
        ws['A1'] = 'Date'
        ws['B1'] = 'Stream'
        ws['C1'] = 'Post Text Preview (500 chars)'
        ws['D1'] = 'Full Post Text'
        ws['E1'] = 'Source URL'
        ws['F1'] = 'Word Count'
        ws['G1'] = 'Format'
        ws['H1'] = 'Notes'
        
        # Set column widths for readability
        ws.column_dimensions['A'].width = 12
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 50
        ws.column_dimensions['D'].width = 70
        ws.column_dimensions['E'].width = 40
        ws.column_dimensions['F'].width = 12
        ws.column_dimensions['G'].width = 15
        ws.column_dimensions['H'].width = 30
        
        next_row = 2
    
    # Append each post as a new row (2 per day in 2-post model)
    for post in posts:
        ws[f'A{next_row}'] = today
        ws[f'B{next_row}'] = post['stream_name']
        ws[f'C{next_row}'] = post['post_text']
        ws[f'D{next_row}'] = post['full_text']
        ws[f'E{next_row}'] = post['source_url']
        ws[f'F{next_row}'] = post['word_count']
        ws[f'G{next_row}'] = post['format_used']
        ws[f'H{next_row}'] = ""  # Notes column for manual annotations
        
        next_row += 1
    
    # Save workbook
    wb.save(output_file)
    print(f"✓ Updated {output_file} with {len(posts)} posts from {today}")
    print(f"  File has {next_row - 2} total post records (Excel row {next_row - 1} is last entry)")
    
    return output_file

def main():
    # Find today's posts file
    today_compact = datetime.date.today().strftime('%Y%m%d')
    posts_file = f'./linkedin_posts_{today_compact}.txt'
    
    if not os.path.exists(posts_file):
        print(f"Error: Posts file not found: {posts_file}")
        print("Make sure you have run the content generation scripts first.")
        return
    
    # Extract posts
    posts = extract_posts_from_file(posts_file)
    
    if not posts:
        print(f"Error: No posts found in {posts_file}")
        return
    
    print(f"✓ Extracted {len(posts)} posts from {posts_file}")
    for post in posts:
        print(f"  - Stream {post['stream_num']}: {post['stream_name']} ({post['word_count']} words)")
    
    # Update Excel
    try:
        output_file = update_excel_history(posts)
        print(f"\n✓ Post history updated: {output_file}")
        print(f"✓ All posts stored locally (no cloud upload)")
    except ImportError as e:
        print(f"\nError: openpyxl not installed. Install with: pip install openpyxl")
        return
    except Exception as e:
        print(f"Error updating Excel file: {e}")
        return

if __name__ == '__main__':
    main()

